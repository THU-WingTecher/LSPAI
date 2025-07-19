import os
import glob
import subprocess
from typing import Dict, List, Tuple
from pathlib import Path
import re
import concurrent.futures
from dataclasses import dataclass
from collections import defaultdict

@dataclass
class TestResult:
    """Data class to store test results"""
    project: str
    model: str
    baseline: str
    target_type: str
    folder_name: str
    coverage_output: float = None
    validrate_output: float = None
    error: str = None

class ParallelRunner:
    """Parallel test runner that extends the original Runner"""
    
    def __init__(self, max_workers: int = 4):
        self.runner = Runner()
        self.max_workers = max_workers
    
    def run_single_test(self, project: str, folder: Path, target_type: str) -> TestResult:
        """Run a single test and return TestResult"""
        try:
            result = self.runner.run_test(project, str(folder))
            
            return TestResult(
                project=project,
                model=self._extract_model_from_path(folder),
                baseline=self._extract_baseline_from_path(folder),
                target_type=target_type,
                folder_name=folder.name,
                coverage_output=result.get("coverage_output"),
                validrate_output=result.get("validrate_output")
            )
        except Exception as e:
            return TestResult(
                project=project,
                model=self._extract_model_from_path(folder),
                baseline=self._extract_baseline_from_path(folder),
                target_type=target_type,
                folder_name=folder.name,
                error=str(e)
            )
    
    def _extract_model_from_path(self, folder: Path) -> str:
        """Extract model name from folder path"""
        path_parts = folder.parts
        models = ['gpt-4o', 'gpt-4o-mini', 'deepseek-chat']
        
        for part in path_parts:
            if part in models:
                return part
            for model in models:
                if f"_{model}_" in part or part.startswith(f"{model}_") or part.endswith(f"_{model}") or part == model:
                    return model
        return "unknown"
    
    def _extract_baseline_from_path(self, folder: Path) -> str:
        """Extract baseline name from folder path"""
        path_parts = folder.parts
        baselines = ['lsprag', 'code_qa', 'naive', 'standard', 'symprompt', 'draco']
        
        for part in path_parts:
            if part in baselines:
                return part
        return "unknown"
    
    def run_tests_parallel(self, organized_folders: Dict[Tuple[str, str, str], Dict[str, List[Path]]]) -> List[TestResult]:
        """Run all tests in parallel"""
        all_tasks = []
        # import pdb; pdb.set_trace()
        # Prepare all tasks
        for (project, model, baseline), target_dict in organized_folders.items():
            for target_type, folders in target_dict.items():
                if folders:
                    for folder in folders:
                        all_tasks.append((project, folder, target_type))
        
        print(f"Running {len(all_tasks)} tests in parallel with {self.max_workers} workers...")
        
        # Execute tasks in parallel
        results = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all tasks
            future_to_task = {
                executor.submit(self.run_single_test, project, folder, target_type): (project, folder, target_type)
                for project, folder, target_type in all_tasks
            }
            
            # Collect results as they complete
            for future in concurrent.futures.as_completed(future_to_task):
                project, folder, target_type = future_to_task[future]
                try:
                    result = future.result()
                    results.append(result)
                    print(f"Completed: {project} + {result.model} + {result.baseline} + {target_type} - {folder.name}")
                except Exception as e:
                    print(f"Task failed for {project} + {folder.name}: {e}")
        
        return results

class ResultSummarizer:
    """Summarize and display test results"""
    
    @staticmethod
    def organize_results(results: List[TestResult]) -> Dict[Tuple[str, str, str], Dict[str, List[TestResult]]]:
        """Organize results by project+model+baseline and target_type"""
        organized = defaultdict(lambda: defaultdict(list))
        
        for result in results:
            key = (result.project, result.model, result.baseline)
            organized[key][result.target_type].append(result)
        
        return organized
    
    @staticmethod
    def print_final_summary(organized_results: Dict[Tuple[str, str, str], Dict[str, List[TestResult]]]):
        """Print final summary showing all 5 results per baseline-model"""
        print("\n" + "="*100)
        print("FINAL RESULTS SUMMARY")
        print("="*100)
        
        for (project, model, baseline), target_dict in organized_results.items():
            print(f"\n{project} + {model} + {baseline}")
            print("-" * 80)
            
            for target_type, results in target_dict.items():
                if results:
                    print(f"\n{target_type.upper()} ({len(results)}/5 results):")
                    
                    # Sort results by folder name for consistent display
                    results.sort(key=lambda x: x.folder_name)
                    
                    for result in results:
                        if result.error:
                            print(f"  {result.folder_name}: ERROR - {result.error}")
                        else:
                            # coverage_debug = result.coverage
                            # validrate_debug = result.validrate
                            coverage_str = f"{result.coverage_output:.4f}" if result.coverage_output is not None else "N/A"
                            validrate_str = f"{result.validrate_output:.4f}" if result.validrate_output is not None else "N/A"
                            print(f"  {result.folder_name}: Coverage={coverage_str}  ValidRate={validrate_str} ")
                    
                    # Calculate averages for successful results
                    successful_results = [r for r in results if r.error is None]
                    
                    if successful_results:
                        coverage_values = [r.coverage_output for r in successful_results if r.coverage_output is not None]
                        validrate_values = [r.validrate_output for r in successful_results if r.validrate_output is not None]
                        
                        if coverage_values:
                            avg_coverage = sum(coverage_values) / len(coverage_values)
                            print(f"  Average Coverage: {avg_coverage:.4f} ({len(coverage_values)}/{len(results)} data points)")
                        
                        if validrate_values:
                            avg_validrate = sum(validrate_values) / len(validrate_values)
                            print(f"  Average Valid Rate: {avg_validrate:.4f} ({len(validrate_values)}/{len(results)} data points)")
                    else:
                        print(f"  No successful results for {target_type}")

    def print_as_csv(self, organized_results: Dict[Tuple[str, str, str], Dict[str, List[TestResult]]]) -> None:
        """
        Print the results as a csv file and also save to CSV files
        project codeQA	StandardRAG	Naive	SymPrompt	LSPRAG	DraCo	LSPRAG-nofix
        cli-4o-mini	0.120916031	0.071653944	0.056386768	0.02778626	0.332926209	None	0.271043257
        cli-4o	0.123256997	0.035012723	0.137709924	0.045903308	0.346870229	None	0.231552163
        cli-deepseek	0.226666667	0.176793893	0.097811705	0.069720102	0.377201018	None	0.287735369
        """
        # First print to console as before
        self._print_csv_to_console(organized_results)
        
        # Then save to CSV files
        self._save_csv_to_files(organized_results)
    
    def _print_csv_to_console(self, organized_results: Dict[Tuple[str, str, str], Dict[str, List[TestResult]]]) -> None:
        """Print CSV format results to console"""
        print("\n" + "="*100)
        print("COVERAGE RESULTS SUMMARY (CSV FORMAT)")
        print("="*100)
        
        # Define baseline order for consistent output
        baseline_order = ['code_qa', 'standard', 'naive', 'symprompt', 'lsprag', 'draco', 'lsprag-nofix']
        baseline_display_names = {
            'code_qa': 'codeQA',
            'standard': 'StandardRAG', 
            'naive': 'Naive',
            'symprompt': 'SymPrompt',
            'lsprag': 'LSPRAG',
            'draco': 'DraCo',
            'lsprag-nofix': 'LSPRAG-nofix'
        }
        
        # Collect all unique project-model combinations
        project_model_combinations = set()
        for (project, model, baseline) in organized_results.keys():
            project_model_combinations.add((project, model))
        
        # Sort project-model combinations for consistent output
        # Order models as: 4o-mini -> 4o -> deepseek
        def model_sort_key(project_model):
            project, model = project_model
            model_order = {'gpt-4o-mini': 0, 'gpt-4o': 1, 'deepseek-chat': 2}
            return (project, model_order.get(model, 999))
        
        project_model_combinations = sorted(project_model_combinations, key=model_sort_key)
        
        # Print header
        header = "project"
        for baseline in baseline_order:
            header += f"\t{baseline_display_names[baseline]}"
        print(header)
        
        # Print results for each project-model combination
        for project, model in project_model_combinations:
            # Create project-model identifier (e.g., "cli-4o-mini")
            project_short = project.replace('commons-', '')  # Remove "commons-" prefix
            model_short = model.replace('gpt-4o-mini', '4o-mini').replace('gpt-4o', '4o').replace('deepseek-chat', 'deepseek')
            row_id = f"{project_short}-{model_short}"
            
            row = row_id
            
            # Get coverage values for each baseline
            for baseline in baseline_order:
                if baseline == 'lsprag-nofix':
                    # For lsprag-nofix, get initial results from lsprag baseline
                    key = (project, model, 'lsprag')
                    if key in organized_results:
                        target_dict = organized_results[key]
                        results = target_dict.get('initial', [])
                        
                        if results:
                            # Calculate average coverage from successful results
                            successful_results = [r for r in results if r.error is None and r.coverage_output is not None]
                            if successful_results:
                                avg_coverage = sum(r.coverage_output for r in successful_results) / len(successful_results)
                                row += f"\t{avg_coverage:.9f}"
                            else:
                                row += "\tNone"
                        else:
                            row += "\tNone"
                    else:
                        row += "\tNone"
                else:
                    # For other baselines (naive, symprompt, lsprag, draco), use final results
                    key = (project, model, baseline)
                    if key in organized_results:
                        target_dict = organized_results[key]
                        results = target_dict.get('final', []) or target_dict.get('codes', [])
                        
                        if results:
                            # Calculate average coverage from successful results
                            successful_results = [r for r in results if r.error is None and r.coverage_output is not None]
                            if successful_results:
                                avg_coverage = sum(r.coverage_output for r in successful_results) / len(successful_results)
                                row += f"\t{avg_coverage:.9f}"
                            else:
                                row += "\tNone"
                        else:
                            row += "\tNone"
                    else:
                        row += "\tNone"
            
            print(row)
        
        print("\n" + "="*100)
        print("VALID RATE RESULTS SUMMARY (CSV FORMAT)")
        print("="*100)
        
        # Print header for valid rate
        print(header)
        
        # Print valid rate results for each project-model combination
        for project, model in project_model_combinations:
            # Create project-model identifier (e.g., "cli-4o-mini")
            project_short = project.replace('commons-', '')  # Remove "commons-" prefix
            model_short = model.replace('gpt-4o-mini', '4o-mini').replace('gpt-4o', '4o').replace('deepseek-chat', 'deepseek')
            row_id = f"{project_short}-{model_short}"
            
            row = row_id
            
            # Get valid rate values for each baseline
            for baseline in baseline_order:
                if baseline == 'lsprag-nofix':
                    # For lsprag-nofix, get initial results from lsprag baseline
                    key = (project, model, 'lsprag')
                    if key in organized_results:
                        target_dict = organized_results[key]
                        results = target_dict.get('initial', [])
                        
                        if results:
                            # Calculate average valid rate from successful results
                            successful_results = [r for r in results if r.error is None and r.validrate_output is not None]
                            if successful_results:
                                avg_validrate = sum(r.validrate_output for r in successful_results) / len(successful_results)
                                row += f"\t{avg_validrate:.9f}"
                            else:
                                row += "\tNone"
                        else:
                            row += "\tNone"
                    else:
                        row += "\tNone"
                else:
                    # For other baselines (naive, symprompt, lsprag, draco), use final results
                    key = (project, model, baseline)
                    if key in organized_results:
                        target_dict = organized_results[key]
                        results = target_dict.get('final', []) or target_dict.get('codes', [])
                        
                        if results:
                            # Calculate average valid rate from successful results
                            successful_results = [r for r in results if r.error is None and r.validrate_output is not None]
                            if successful_results:
                                avg_validrate = sum(r.validrate_output for r in successful_results) / len(successful_results)
                                row += f"\t{avg_validrate:.9f}"
                            else:
                                row += "\tNone"
                        else:
                            row += "\tNone"
                    else:
                        row += "\tNone"
            
            print(row)
    
    def _save_csv_to_files(self, organized_results: Dict[Tuple[str, str, str], Dict[str, List[TestResult]]]) -> None:
        """Save CSV and Excel format results to files"""
        import csv
        from datetime import datetime
        
        # Try to import openpyxl for Excel support
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            excel_available = True
        except ImportError:
            excel_available = False
            print("Warning: openpyxl not installed. Excel files will not be generated.")
            print("Install with: pip install openpyxl")
        
        # Define baseline order for consistent output
        baseline_order = ['code_qa', 'standard', 'naive', 'symprompt', 'lsprag', 'draco', 'lsprag-nofix']
        baseline_display_names = {
            'code_qa': 'codeQA',
            'standard': 'StandardRAG', 
            'naive': 'Naive',
            'symprompt': 'SymPrompt',
            'lsprag': 'LSPRAG',
            'draco': 'DraCo',
            'lsprag-nofix': 'LSPRAG-nofix'
        }
        
        # Collect all unique project-model combinations
        project_model_combinations = set()
        for (project, model, baseline) in organized_results.keys():
            project_model_combinations.add((project, model))
        
        # Sort project-model combinations for consistent output
        def model_sort_key(project_model):
            project, model = project_model
            model_order = {'gpt-4o-mini': 0, 'gpt-4o': 1, 'deepseek-chat': 2}
            return (project, model_order.get(model, 999))
        
        project_model_combinations = sorted(project_model_combinations, key=model_sort_key)
        
        # Create timestamp for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Prepare data for both CSV and Excel
        coverage_data = []
        validrate_data = []
        
        # Prepare coverage data
        for project, model in project_model_combinations:
            project_short = project.replace('commons-', '')
            model_short = model.replace('gpt-4o-mini', '4o-mini').replace('gpt-4o', '4o').replace('deepseek-chat', 'deepseek')
            row_id = f"{project_short}-{model_short}"
            
            row = [row_id]
            
            # Get coverage values for each baseline
            for baseline in baseline_order:
                if baseline == 'lsprag-nofix':
                    key = (project, model, 'lsprag')
                    if key in organized_results:
                        target_dict = organized_results[key]
                        results = target_dict.get('initial', [])
                        
                        if results:
                            successful_results = [r for r in results if r.error is None and r.coverage_output is not None]
                            if successful_results:
                                avg_coverage = sum(r.coverage_output for r in successful_results) / len(successful_results)
                                row.append(avg_coverage)
                            else:
                                row.append(None)
                        else:
                            row.append(None)
                    else:
                        row.append(None)
                else:
                    key = (project, model, baseline)
                    if key in organized_results:
                        target_dict = organized_results[key]
                        results = target_dict.get('final', []) or target_dict.get('codes', [])
                        
                        if results:
                            successful_results = [r for r in results if r.error is None and r.coverage_output is not None]
                            if successful_results:
                                avg_coverage = sum(r.coverage_output for r in successful_results) / len(successful_results)
                                row.append(avg_coverage)
                            else:
                                row.append(None)
                        else:
                            row.append(None)
                    else:
                        row.append(None)
            
            coverage_data.append(row)
        
        # Prepare valid rate data
        for project, model in project_model_combinations:
            project_short = project.replace('commons-', '')
            model_short = model.replace('gpt-4o-mini', '4o-mini').replace('gpt-4o', '4o').replace('deepseek-chat', 'deepseek')
            row_id = f"{project_short}-{model_short}"
            
            row = [row_id]
            
            # Get valid rate values for each baseline
            for baseline in baseline_order:
                if baseline == 'lsprag-nofix':
                    key = (project, model, 'lsprag')
                    if key in organized_results:
                        target_dict = organized_results[key]
                        results = target_dict.get('initial', [])
                        
                        if results:
                            successful_results = [r for r in results if r.error is None and r.validrate_output is not None]
                            if successful_results:
                                avg_validrate = sum(r.validrate_output for r in successful_results) / len(successful_results)
                                row.append(avg_validrate)
                            else:
                                row.append(None)
                        else:
                            row.append(None)
                    else:
                        row.append(None)
                else:
                    key = (project, model, baseline)
                    if key in organized_results:
                        target_dict = organized_results[key]
                        results = target_dict.get('final', []) or target_dict.get('codes', [])
                        
                        if results:
                            successful_results = [r for r in results if r.error is None and r.validrate_output is not None]
                            if successful_results:
                                avg_validrate = sum(r.validrate_output for r in successful_results) / len(successful_results)
                                row.append(avg_validrate)
                            else:
                                row.append(None)
                        else:
                            row.append(None)
                    else:
                        row.append(None)
            
            validrate_data.append(row)
        
        # Save CSV files
        coverage_filename = f"coverage_results_{timestamp}.csv"
        with open(coverage_filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            header = ['project'] + [baseline_display_names[baseline] for baseline in baseline_order]
            writer.writerow(header)
            
            # Write data rows
            for row in coverage_data:
                csv_row = [row[0]] + [f"{val:.9f}" if val is not None else "None" for val in row[1:]]
                writer.writerow(csv_row)
        
        validrate_filename = f"validrate_results_{timestamp}.csv"
        with open(validrate_filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            header = ['project'] + [baseline_display_names[baseline] for baseline in baseline_order]
            writer.writerow(header)
            
            # Write data rows
            for row in validrate_data:
                csv_row = [row[0]] + [f"{val:.9f}" if val is not None else "None" for val in row[1:]]
                writer.writerow(csv_row)
        
        # Save Excel file if openpyxl is available
        excel_filename = None
        if excel_available:
            excel_filename = f"test_results_{timestamp}.xlsx"
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create Coverage sheet
            ws_coverage = wb.create_sheet("Coverage")
            ws_coverage.title = "Coverage"
            
            # Add header to coverage sheet
            header = ['project'] + [baseline_display_names[baseline] for baseline in baseline_order]
            for col, value in enumerate(header, 1):
                cell = ws_coverage.cell(row=1, column=col, value=value)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data to coverage sheet
            for row_idx, row_data in enumerate(coverage_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_coverage.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 1:  # Project column
                        cell.font = Font(bold=True)
                    elif value is not None:  # Numeric values
                        cell.number_format = '0.000000000'
            
            # Create Valid Rate sheet
            ws_validrate = wb.create_sheet("Valid Rate")
            ws_validrate.title = "Valid Rate"
            
            # Add header to valid rate sheet
            for col, value in enumerate(header, 1):
                cell = ws_validrate.cell(row=1, column=col, value=value)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data to valid rate sheet
            for row_idx, row_data in enumerate(validrate_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_validrate.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 1:  # Project column
                        cell.font = Font(bold=True)
                    elif value is not None:  # Numeric values
                        cell.number_format = '0.000000000'
            
            # Auto-adjust column widths
            for ws in [ws_coverage, ws_validrate]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save Excel file
            wb.save(excel_filename)
        
        # Print file information
        print(f"\nFiles saved:")
        print(f"  Coverage results: {coverage_filename}")
        print(f"  Valid rate results: {validrate_filename}")
        if excel_filename:
            print(f"  Excel results: {excel_filename}")

class Parser:
    """
    This parser parse the printed output of the test result. 
    we return a dict with following keys:
        {
            "validrate": 4 / 24,
            "coverage": 94 / 431,
        }
    """
    @staticmethod
    def go_output_parser(output: str) -> dict:
        """
        -------------------
        Files: 4/24 passed (16.67%)
        Functions: 18/18 passed (100.00%)
        -------------------
        -------------------
        Coverage Report: /LSPAI/experiments/projects/logrus/lspai-workspace/7_9_2025__08_44_27/logrus/cfg_withcontext_original/gpt-4o/results/initial-report/coverage.out
        Total Statements: 431
        Covered Statements: 94
        Coverage Percentage: 21.81%
        -------------------

        from above printed output, we get following information. 
        validrate = 4 / 24  
        coverage = 94 / 431 

        """
        result = {}
        
        # Extract valid rate information from "Files: X/Y passed" line
        files_match = re.search(r'Files: (\d+)/(\d+) passed', output)
        
        if files_match:
            passed_files = int(files_match.group(1))
            total_files = int(files_match.group(2))
            result["validrate"] = f"{passed_files} / {total_files}"
            result["validrate_output"] = passed_files / total_files if total_files > 0 else 0.0
        
        # Extract coverage information from "Total Statements" and "Covered Statements" lines
        total_statements_match = re.search(r'Total Statements: (\d+)', output)
        covered_statements_match = re.search(r'Covered Statements: (\d+)', output)
        
        if total_statements_match and covered_statements_match:
            total_statements = int(total_statements_match.group(1))
            covered_statements = int(covered_statements_match.group(1))
            result["coverage"] = f"{covered_statements} / {total_statements}"
            result["coverage_output"] = covered_statements / total_statements if total_statements > 0 else 0.0
        
        return result
    
    @staticmethod
    def java_output_parser(output: str) -> dict:
        """
============================
Total lines 1965
Missed Lines 1252
Line Coverages are 36.28%
============================
Printing valid rate
Pass rate for /LSPAI/experiments/data/commons-cli/lsprag/1/gpt-4o/results/final
============================
Total .java files: 44
Files with corresponding .class files: 24
Pass rate: 54.55%
============================
        """
        result = {}
        
        # Extract coverage information
        # Pattern to match "Total lines X" and "Missed Lines Y"
        total_lines_match = re.search(r'Total lines (\d+)', output)
        missed_lines_match = re.search(r'Missed Lines (\d+)', output)
        
        if total_lines_match and missed_lines_match:
            total_lines = int(total_lines_match.group(1))
            missed_lines = int(missed_lines_match.group(1))
            covered_lines = total_lines - missed_lines
            result["coverage"] = f"{covered_lines} / {total_lines}"
            result["coverage_output"] = covered_lines / total_lines if total_lines > 0 else 0.0
        
        # Extract valid rate information
        # Pattern to match "Total .java files: X" and "Files with corresponding .class files: Y"
        total_java_files_match = re.search(r'Total \.java files: (\d+)', output)
        class_files_match = re.search(r'Files with corresponding \.class files: (\d+)', output)
        
        if total_java_files_match and class_files_match:
            total_java_files = int(total_java_files_match.group(1))
            class_files = int(class_files_match.group(1))
            result["validrate"] = f"{class_files} / {total_java_files}"
            result["validrate_output"] = class_files / total_java_files if total_java_files > 0 else 0.0
        
        return result
    
    @staticmethod
    def python_output_parser(output: str, project_name: str = None) -> dict:
        """
------------------------------------------------------
TOTAL                               7405   3413    54%
Coverage collection completed. Summary saved to /LSPAI/experiments/data/main_result/black/code_qa/codeQA_deepseek-chat_20250531_112041/codes-report/summary.txt
PassRate ((passed files + failed files)/ total files): 201/293
        """
        result = {}
        
        # Define total lines for each Python project
        project_total_lines = {
            "black": 7523,
            "tornado": 8955,  # Update this with actual value for tornado
        }
        
        # Extract coverage information
        # Pattern to match "TOTAL X Y Z%" where X=missed lines, Y=covered lines, Z=percentage
        coverage_match = re.search(r'TOTAL\s+(\d+)\s+(\d+)\s+(\d+)%', output)
        
        if coverage_match:
            missed_lines = int(coverage_match.group(1))  # 7405
            covered_lines_from_output = int(coverage_match.group(2))  # 3413
            
            # Calculate actual covered lines: missed_lines - covered_lines_from_output
            actual_covered_lines = missed_lines - covered_lines_from_output  # 7405 - 3413 = 3992
            
            # Use project-specific total lines
            total_lines = project_total_lines.get(project_name, 7555)  # Default to 7555 if project not found
            
            result["coverage"] = f"{actual_covered_lines} / {total_lines}"
            result["coverage_output"] = actual_covered_lines / total_lines if total_lines > 0 else 0.0
        
        # Extract valid rate information
        # Pattern to match "PassRate ((passed files + failed files)/ total files): X/Y"
        validrate_match = re.search(r'PassRate.*?: (\d+)/(\d+)', output)
        
        if validrate_match:
            passed_files = int(validrate_match.group(1))
            total_files = int(validrate_match.group(2))
            result["validrate"] = f"{passed_files} / {total_files}"
            result["validrate_output"] = passed_files / total_files if total_files > 0 else 0.0
        
        return result


class Runner: 
    """ 
    Run the test and get the result
    run command : 
    projects: commons-cli, commons-csv ==> bash /LSPAI/experiments/scripts/java_coverage.bash <project_path> <experiment_save_folder_path> 
    projects: black, tornado ==> bash /LSPAI/experiments/scripts/go_coverage.bash <project_path> <experiment_save_folder_path> 
    projects: cobra, logrus ==> bash /LSPAI/experiments/scripts/python_coverage.bash <project_path> <experiment_save_folder_path> 
    """
    projectpath = {
        "commons-cli": "/LSPAI/experiments/projects/commons-cli",
        "commons-csv": "/LSPAI/experiments/projects/commons-csv",
        "black": "/LSPAI/experiments/projects/black",
        "tornado": "/LSPAI/experiments/projects/tornado",
        "cobra": "/LSPAI/experiments/projects/cobra",
        "logrus": "/LSPAI/experiments/projects/logrus",
    }
    
    def __init__(self):
        # Define which script to use for each project
        self.project_scripts = {
            "commons-cli": "/LSPAI/scripts/java_coverage.bash",
            "commons-csv": "/LSPAI/scripts/java_coverage.bash",
            "black": "/LSPAI/scripts/python_coverage.bash",
            "tornado": "/LSPAI/scripts/python_coverage.bash",
            "cobra": "/LSPAI/scripts/go_coverage.bash",
            "logrus": "/LSPAI/scripts/go_coverage.bash",
        }
        
        # Define which parser to use for each project
        self.project_parsers = {
            "commons-cli": Parser.java_output_parser,
            "commons-csv": Parser.java_output_parser,
            "black": Parser.python_output_parser,
            "tornado": Parser.python_output_parser,
            "cobra": Parser.go_output_parser,
            "logrus": Parser.go_output_parser,
        }
    
    def run_test(self, project_name: str, experiment_save_folder_path: str) -> dict:
        """
        Run the test for a specific project and experiment folder
        
        Args:
            project_name: Name of the project (e.g., "commons-cli", "black", "cobra")
            experiment_save_folder_path: Path to the experiment folder
            
        Returns:
            Dictionary with parsed test results containing "validrate" and "coverage"
        """
        if project_name not in self.projectpath:
            raise ValueError(f"Unknown project: {project_name}")
        
        if project_name not in self.project_scripts:
            raise ValueError(f"No script defined for project: {project_name}")
        
        # Get the project path and script path
        project_path = self.projectpath[project_name]
        script_path = self.project_scripts[project_name]
        
        # Check if paths exist
        if not os.path.exists(project_path):
            raise FileNotFoundError(f"Project path does not exist: {project_path}")
        
        if not os.path.exists(script_path):
            raise FileNotFoundError(f"Script does not exist: {script_path}")
        
        if not os.path.exists(experiment_save_folder_path):
            raise FileNotFoundError(f"Experiment folder does not exist: {experiment_save_folder_path}")
        
        # Run the test command
        try:
            if project_name in ["black", "tornado"]:
                # Both black and tornado need conda environment
                conda_env_name = project_name  # Use project name as conda env name
                cmd = ["conda", "run", "-n", conda_env_name, "bash", script_path, project_path, experiment_save_folder_path]
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    check=True,
                    cwd=os.path.dirname(script_path)
                )
            else :
                cmd = ["bash", script_path, project_path, experiment_save_folder_path]
                result = subprocess.run(
                    cmd,
                    capture_output=True,
                    text=True,
                    check=True,
                    cwd=os.path.dirname(script_path)
                )
                
            # Get the output
            output = result.stdout
            # Print the bash command output
            print(f"Command output for {project_name}:")
            print("-" * 50)
            print(output)
            print("-" * 50)

            # Parse the output using the appropriate parser
            parser = self.project_parsers[project_name]
            
            if project_name in ["black", "tornado"]:  # Python projects
                parsed_result = parser(output, project_name)
            else:
                parsed_result = parser(output)
            
            return parsed_result
            
        except subprocess.CalledProcessError as e:
            print(f"Error running test for {project_name}: {e}")
            print(f"Command: {' '.join(cmd)}")
            print(f"Error output: {e.stderr}")
            raise
        except Exception as e:
            print(f"Unexpected error running test for {project_name}: {e}")
            raise

class FileFounder:
    """
    We find the data results by finding the folder from data folder. 

    Target Folder: **/final/**, **/initial/**, or **/codes 
    After find out the folder, we classify them by the following rules: 
    1. project name : black, tornado, commons-cli, commons-csv, cobra, logrus. 
    2. model name : gpt-4o, gpt-4o-mini, [deepseek-chat or deepseek] 
    3. baseline name : lsprag, code_qa, naive, standard, symprompt, draco 

    each set should have five folders : (black+gpt-4o+lsprag/**/final folder should have 5 , put assert if there is no 5 folders )
    """
    
    def __init__(self, data_root_path: str):
        """
        Initialize FileFounder with the root path of data folder
        
        Args:
            data_root_path: Root path where the data folders are located
        """
        self.data_root_path = Path(data_root_path)
        self.projects = ['black', 'tornado', 'commons-cli', 'commons-csv', 'cobra', 'logrus']
        self.models = ['gpt-4o', 'gpt-4o-mini', 'deepseek-chat']
        self.baselines = ['lsprag', 'code_qa', 'naive', 'standard', 'symprompt', 'draco']
        self.target_folders = ['final', 'initial', 'codes']
        
    def find_target_folders(self) -> List[Path]:
        """
        Find target folders based on baseline type:
        - For "lsprag" baseline: save both "initial" and "final" folders
        - For other baselines (naive, symprompt, etc.): save only "final" folders
        
        Returns:
            List of Path objects pointing to target folders
        """
        target_folders = []
        
        # First, find all potential target folders
        all_potential_folders = []
        for target in self.target_folders:
            pattern = str(self.data_root_path / "**" / target)
            found_folders = glob.glob(pattern, recursive=True)
            all_potential_folders.extend([Path(folder) for folder in found_folders if Path(folder).is_dir()])
        
        # Filter folders based on baseline type
        for folder in all_potential_folders:
            project, model, baseline, target_type = self.classify_folder(folder)
            
            # Only include folders that we can properly classify
            if project and model and baseline and target_type:
                # For lsprag baseline, include both initial and final
                if baseline == "lsprag":
                    if target_type in ["initial", "final"]:
                        target_folders.append(folder)
                # For other baselines, only include final
                else:
                    if target_type in ["final", "codes"]:
                        target_folders.append(folder)
        
        return target_folders
    
    def classify_folder(self, folder_path: Path) -> Tuple[str, str, str, str]:
        """
        Classify a folder based on its path structure
        
        Args:
            folder_path: Path to the folder to classify
            
        Returns:
            Tuple of (project_name, model_name, baseline_name, target_type)
        """
        path_parts = folder_path.parts
        
        # Find project name
        project_name = None
        for part in path_parts:
            if part in self.projects:
                project_name = part
                break
        
        # Find model name - handle embedded model names in folder names
        model_name = None
        for part in path_parts:
            # First check for exact matches
            if part in self.models:
                model_name = part
                break
            # Then check if model name is embedded in folder name (e.g., "codeQA_gpt-4o-mini_20250707_143750")
            # Use word boundaries to avoid partial matches
            for model in self.models:
                # Check if model appears as a complete word in the folder name
                if f"_{model}_" in part or part.startswith(f"{model}_") or part.endswith(f"_{model}") or part == model:
                    model_name = model
                    break
            if model_name:
                break
        
        # Find baseline name
        baseline_name = None
        for part in path_parts:
            if part in self.baselines:
                baseline_name = part
                break
        
        # Find target type (final, initial, codes)
        target_type = None
        for part in path_parts:
            if part in self.target_folders:
                target_type = part
                break
        
        return project_name, model_name, baseline_name, target_type
    
    def organize_folders(self) -> Dict[Tuple[str, str, str], Dict[str, List[Path]]]:
        """
        Organize all found folders by project, model, and baseline
        
        Returns:
            Dictionary with structure: {(project, model, baseline): {target_type: [folders]}}
        """
        target_folders = self.find_target_folders()
        organized = {}
        
        for folder in target_folders:
            project, model, baseline, target_type = self.classify_folder(folder)
            
            if project and model and baseline and target_type:
                key = (project, model, baseline)
                if key not in organized:
                    organized[key] = {target: [] for target in self.target_folders}
                
                organized[key][target_type].append(folder)
        
        return organized
    
    def validate_folder_counts(self, organized_folders: Dict[Tuple[str, str, str], Dict[str, List[Path]]]) -> None:
        """
        Validate that each combination has exactly 5 folders for each target type
        
        Args:
            organized_folders: Organized folder structure from organize_folders()
            
        Raises:
            AssertionError: If any combination doesn't have exactly 5 folders
        """
        for (project, model, baseline), target_dict in organized_folders.items():
            for target_type, folders in target_dict.items():
                if folders:  # Only check if there are folders for this target type
                    if len(folders) != 5:
                        print(f"Expected 5 folders for {project}+{model}+{baseline}/{target_type}, found {len(folders)}")
                        # assert len(folders) == 5, f"Expected 5 folders for {project}+{model}+{baseline}/{target_type}, found {len(folders)}"
    
    def get_folders_for_combination(self, project: str, model: str, baseline: str, target_type: str = None) -> List[Path]:
        """
        Get folders for a specific combination of project, model, and baseline
        
        Args:
            project: Project name
            model: Model name
            baseline: Baseline name
            target_type: Optional target type filter (final, initial, codes)
            
        Returns:
            List of Path objects for the specified combination
        """
        organized = self.organize_folders()
        key = (project, model, baseline)
        
        if key not in organized:
            return []
        
        if target_type:
            return organized[key].get(target_type, [])
        else:
            all_folders = []
            for folders in organized[key].values():
                all_folders.extend(folders)
            return all_folders
    
    def run_validation(self) -> Dict[Tuple[str, str, str], Dict[str, int]]:
        """
        Run full validation and return folder counts for each combination
        
        Returns:
            Dictionary with folder counts for each combination and target type
        """
        organized = self.organize_folders()
        self.validate_folder_counts(organized)
        
        # Return summary of folder counts
        summary = {}
        for (project, model, baseline), target_dict in organized.items():
            summary[(project, model, baseline)] = {
                target_type: len(folders) for target_type, folders in target_dict.items()
            }
        
        return summary
    def print_found_folders(self) -> None:
        """
        Print out all found target folders with their full paths for debugging
        """
        target_folders = self.find_target_folders()
        print(f"Found {len(target_folders)} target folders:")
        print("=" * 80)
        
        for i, folder in enumerate(target_folders, 1):
            project, model, baseline, target_type = self.classify_folder(folder)
            print(f"{i:3d}. {folder}")
            print(f"     Project: {project}, Model: {model}, Baseline: {baseline}, Type: {target_type}")
            print()
        
        print("=" * 80)
        print(f"Total folders found: {len(target_folders)}")

    def print_organized_folders(self) -> None:
        """
        Print out organized folders by project+model+baseline combination
        """
        organized = self.organize_folders()
        print("Organized folders by project+model+baseline:")
        print("=" * 80)
        
        for (project, model, baseline), target_dict in organized.items():
            print(f"\n{project} + {model} + {baseline}:")
            for target_type, folders in target_dict.items():
                if folders:
                    print(f"  {target_type}: {len(folders)} folders")
                    for folder in folders:
                        print(f"    - {folder}")
        
        print("=" * 80)

if __name__ == "__main__":
    # data_root_path = "/LSPAI/experiments/data/main_result/commons-cli"
    import sys 
    data_root_path = sys.argv[1]
    # data_root_path = "/LSPAI/experiments/data/main_result/commons-cli"
    file_founder = FileFounder(data_root_path)
    max_workers = 30
    # Print all found folders first
    print("=== FOUND FOLDERS ===")
    file_founder.print_found_folders()
    
    print("\n=== ORGANIZED FOLDERS ===")
    file_founder.print_organized_folders()
    
    # Then run validation
    print("\n=== VALIDATION RESULTS ===")
    summary = file_founder.run_validation()
    for (project, model, baseline), counts in summary.items():
        print(f"{project} + {model} + {baseline}: {counts}")
    
    print("\n=== TEST RESULTS AND AVERAGES ===")
    organized = file_founder.organize_folders()
    print(organized)
    # Run tests in parallel
    # test_csv_printing_with_mock_data()

    parallel_runner = ParallelRunner(max_workers=max_workers)
    all_results = parallel_runner.run_tests_parallel(organized)
    # Organize and display final results
    summarizer = ResultSummarizer()
    organized_results = summarizer.organize_results(all_results)
    summarizer.print_final_summary(organized_results)
    # Print CSV format results
    summarizer.print_as_csv(organized_results)